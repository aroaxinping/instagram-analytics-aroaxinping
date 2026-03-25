"""
generate_excel.py — Genera el Excel de analytics de Instagram con fórmulas nativas.

Hojas:
  1. 📊 Overview          — KPIs del periodo + fuentes de tráfico
  2. 🎬 Reels Raw         — todos los reels con métricas brutas
  3. 📐 Engagement Calc   — fórmulas interactivas de engagement
  4. 📈 Métricas Avanzadas — métricas de la industria (amplification, quality score…)

Uso:
    python3 src/generate_excel.py
"""

import os, sys
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
from openpyxl.utils import get_column_letter

BASE  = Path(__file__).resolve().parent.parent
PROC  = BASE / "data" / "processed"
RAW   = BASE / "data" / "raw"

TOTAL_FOLLOWERS = 8728   # seguidores totales al cierre del periodo (23 mar 2026)

# ── Paleta ────────────────────────────────────────────────────────────────────
BG_DARK    = "FF0F0F1A"   # fondo oscuro
ACCENT     = "FFB388EB"   # lavanda
ACCENT2    = "FFE6A8D7"   # rosa
ROW1       = "FF1A1A2E"
ROW2       = "FF16213E"
GOLD_C     = "FFFCD34D"
WHITE      = "FFE2E8F0"

def mk_fill(hex6): return PatternFill("solid", fgColor=hex6)

ACCENT_FILL   = mk_fill("FF2D1B69")
LAVENDER_FILL = mk_fill("FF3B0764")
ROW_FILL_1    = mk_fill(ROW1)
ROW_FILL_2    = mk_fill(ROW2)

title_font   = Font(name="Calibri", bold=True,  size=13, color=ACCENT)
subhead_font = Font(name="Calibri", bold=True,  size=11, color=ACCENT2)
header_font  = Font(name="Calibri", bold=True,  size=10, color=GOLD_C)
bold_font    = Font(name="Calibri", bold=True,  size=10, color=WHITE)
normal_font  = Font(name="Calibri", bold=False, size=10, color=WHITE)

def thin_border():
    s = Side(style="thin", color="FF334155")
    return Border(left=s, right=s, top=s, bottom=s)

def style_header(cell):
    cell.font      = header_font
    cell.fill      = ACCENT_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = thin_border()

def style_cell(cell, bold=False, fill=None, num_format=None, align="right"):
    cell.font      = bold_font if bold else normal_font
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
    cell.border    = thin_border()
    if fill:       cell.fill       = fill
    if num_format: cell.number_format = num_format

# ── Carga de datos ────────────────────────────────────────────────────────────
print("📂 Cargando datos procesados...")
reels  = pd.read_csv(PROC / "reels_metricas.csv")
daily  = pd.read_csv(PROC / "metricas_diarias.csv")

# Filtrar solo Reels (descartar Stories si hubiera)
reels = reels[reels["visualizaciones"] > 0].copy()
reels = reels.sort_values("visualizaciones", ascending=False).reset_index(drop=True)

print(f"  {len(reels)} reels · {len(daily)} días de datos")

wb = Workbook()
wb.remove(wb.active)   # eliminar hoja por defecto

# ═════════════════════════════════════════════════════════════════════════════
# SHEET 1: OVERVIEW
# ═════════════════════════════════════════════════════════════════════════════
ws1 = wb.create_sheet("📊 Overview")
ws1.sheet_view.showGridLines = False

ws1.merge_cells("A1:E1")
ws1["A1"] = "📊 Instagram Analytics — aroaxinping | 24 Feb – 23 Mar 2026"
ws1["A1"].font = title_font; ws1["A1"].fill = ACCENT_FILL
ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 35

for w, col in zip([22,18,14,14,34], "ABCDE"):
    ws1.column_dimensions[col].width = w

headers1 = ["Métrica", "Valor", "Δ periodo", "Δ %", "Interpretación"]
for i, h in enumerate(headers1, 1):
    c = ws1.cell(row=3, column=i, value=h); style_header(c)
ws1.row_dimensions[3].height = 22

overview_kpis = [
    ("Período analizado",       "28 días",        None,  None,  "24 feb – 23 mar 2026"),
    ("Reels publicados",        len(reels),        None,  None,  "En el período analizado"),
    ("Visualizaciones totales", int(reels["visualizaciones"].sum()), None, None, "Reproducciones acumuladas"),
    ("Alcance total",           int(reels["alcance"].sum()),         None, None, "Cuentas únicas alcanzadas"),
    ("Likes totales",           int(reels["me_gustas"].sum()),       None, None, ""),
    ("Comentarios",             int(reels["comentarios"].sum()),     None, None, ""),
    ("Guardados",               int(reels["guardados"].sum()),       None, None, "Señal de calidad clave en Instagram"),
    ("Compartidos",             int(reels["compartidos"].sum()),     None, None, ""),
    ("Seguidores ganados",      int(reels["seguidores_ganados"].sum()), None, None, "+X sobre base"),
    ("Seguidores totales",      TOTAL_FOLLOWERS,  None,  None,  "Al cierre del periodo"),
    ("Engagement rate medio",   f"{reels['engagement_rate'].mean():.1f}%", None, None, "Calculado sobre alcance"),
    ("Save rate medio",         f"{reels['save_rate'].mean():.2f}%",       None, None, "Guardados / alcance — señal de valor"),
    ("Share rate medio",        f"{reels['share_rate'].mean():.2f}%",      None, None, "Compartidos / vistas"),
]

for row_idx, (metric, val, chg_abs, chg_pct, interp) in enumerate(overview_kpis, 4):
    fill = ROW_FILL_1 if row_idx % 2 == 0 else ROW_FILL_2
    ws1.cell(row=row_idx, column=1, value=metric).font = bold_font
    ws1.cell(row=row_idx, column=1).fill = fill; ws1.cell(row=row_idx, column=1).border = thin_border()
    ws1.cell(row=row_idx, column=1).alignment = Alignment(horizontal="left")
    c_val = ws1.cell(row=row_idx, column=2, value=val)
    style_cell(c_val, bold=True, fill=fill, num_format="#,##0" if isinstance(val, int) else None)
    ws1.cell(row=row_idx, column=3, value=chg_abs or "—"); style_cell(ws1.cell(row=row_idx, column=3), fill=fill)
    ws1.cell(row=row_idx, column=4, value=chg_pct or "—"); style_cell(ws1.cell(row=row_idx, column=4), fill=fill)
    c_i = ws1.cell(row=row_idx, column=5, value=interp); style_cell(c_i, fill=fill, align="left")
    ws1.row_dimensions[row_idx].height = 22

# Top 5 reels
top_row = len(overview_kpis) + 5
ws1.merge_cells(f"A{top_row}:E{top_row}")
ws1.cell(row=top_row, column=1, value="🏆 TOP 5 REELS POR VISTAS").font = subhead_font
ws1.cell(row=top_row, column=1).fill = ACCENT_FILL
ws1.cell(row=top_row, column=1).alignment = Alignment(horizontal="left")
ws1.row_dimensions[top_row].height = 24

top_row += 1
for i, h in enumerate(["#", "Descripción", "Vistas", "Engagement %", "Guardados"], 1):
    c = ws1.cell(row=top_row, column=i, value=h); style_header(c)

for j, (_, row) in enumerate(reels.head(5).iterrows(), top_row + 1):
    fill = ROW_FILL_1 if j % 2 == 0 else ROW_FILL_2
    for col, val in [(1, j - top_row), (2, row["descripcion_corta"][:45]),
                     (3, int(row["visualizaciones"])), (4, f"{row['engagement_rate']:.1f}%"),
                     (5, int(row["guardados"]))]:
        cell = ws1.cell(row=j, column=col, value=val)
        cell.font = normal_font; cell.fill = fill; cell.border = thin_border()
        cell.alignment = Alignment(horizontal="left" if col == 2 else "right", vertical="center")
        if col == 3: cell.number_format = "#,##0"
    ws1.row_dimensions[j].height = 20

# ═════════════════════════════════════════════════════════════════════════════
# SHEET 2: REELS RAW
# ═════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("🎬 Reels Raw")
ws2.sheet_view.showGridLines = False

col_names2 = ["#", "Fecha", "Franja", "Tema", "Duración\n(s)", "Vistas",
              "Alcance", "Likes", "Coments", "Guardados", "Compartidos",
              "Seg.\nGanados", "ER %", "Save\nRate %", "Share\nRate %", "Descripción"]
col_widths2 = [4,  12,  16,   18,   9,    12,   12,   10,  9,     10,     12,
               9,    8,   9,    9,   40]

ws2.merge_cells(f"A1:{get_column_letter(len(col_names2))}1")
ws2["A1"] = "🎬 Reels — Datos Brutos (24 Feb – 23 Mar 2026)"
ws2["A1"].font = title_font; ws2["A1"].fill = ACCENT_FILL
ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 30

for i, (h, w) in enumerate(zip(col_names2, col_widths2), 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
    c = ws2.cell(row=2, column=i, value=h); style_header(c)
ws2.row_dimensions[2].height = 30

for r_idx, (_, row) in enumerate(reels.iterrows(), 3):
    fill = ROW_FILL_1 if r_idx % 2 == 0 else ROW_FILL_2
    vals = [r_idx - 2, row["fecha"], row["franja"], row["tema"],
            int(row["duracion_seg"]), int(row["visualizaciones"]),
            int(row["alcance"]), int(row["me_gustas"]), int(row["comentarios"]),
            int(row["guardados"]), int(row["compartidos"]), int(row["seguidores_ganados"]),
            round(row["engagement_rate"], 2), round(row["save_rate"], 2),
            round(row["share_rate"], 2), row["descripcion_corta"]]
    for c_idx, val in enumerate(vals, 1):
        cell = ws2.cell(row=r_idx, column=c_idx, value=val)
        cell.font = normal_font; cell.fill = fill; cell.border = thin_border()
        cell.alignment = Alignment(
            horizontal="left" if c_idx in (2,3,4,16) else "right", vertical="center")
        if c_idx in (6,7,8,9,10,11,12): cell.number_format = "#,##0"
        if c_idx in (13,14,15):         cell.number_format = "0.00"
    ws2.row_dimensions[r_idx].height = 18

# ═════════════════════════════════════════════════════════════════════════════
# SHEET 3: ENGAGEMENT CALC (FÓRMULAS NATIVAS)
# ═════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("📐 Engagement Calc")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:L1")
ws3["A1"] = "📐 Calculadora de Engagement — Fórmulas Automáticas (base: alcance)"
ws3["A1"].font = title_font; ws3["A1"].fill = ACCENT_FILL
ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 35

col_names3  = ["#", "Descripción", "Alcance", "Likes", "Coments",
               "Guardados", "Compartidos", "Engagement\nRate %",
               "Like\nRate %", "Save\nRate %", "Share\nRate %", "Follower\nConv %"]
col_widths3 = [4, 36, 12, 10, 10, 10, 12, 14, 12, 12, 12, 13]

for i, (h, w) in enumerate(zip(col_names3, col_widths3), 1):
    ws3.column_dimensions[get_column_letter(i)].width = w
    c = ws3.cell(row=2, column=i, value=h); style_header(c)
ws3.row_dimensions[2].height = 32

# seg_ganados not in this sheet but needed for follower conv — store separately
seg_ganados_vals = []

for r_idx, (_, row) in enumerate(reels.iterrows(), 3):
    fill = ROW_FILL_1 if r_idx % 2 == 0 else ROW_FILL_2
    R = r_idx
    seg_ganados_vals.append(int(row["seguidores_ganados"]))

    raw_vals = [(1, r_idx-2), (2, row["descripcion_corta"][:40]),
                (3, int(row["alcance"])), (4, int(row["me_gustas"])),
                (5, int(row["comentarios"])), (6, int(row["guardados"])),
                (7, int(row["compartidos"]))]
    for col, val in raw_vals:
        cell = ws3.cell(row=R, column=col, value=val)
        cell.font = normal_font; cell.fill = fill; cell.border = thin_border()
        cell.alignment = Alignment(horizontal="left" if col==2 else "right", vertical="center")
        if col >= 3: cell.number_format = "#,##0"

    # Col H: Engagement Rate = (likes+comments+saves+shares) / alcance * 100
    c = ws3.cell(row=R, column=8)
    c.value = f"=IF(C{R}=0,0,(D{R}+E{R}+F{R}+G{R})/C{R}*100)"
    c.font = bold_font; c.fill = LAVENDER_FILL; c.border = thin_border()
    c.number_format = "0.00"; c.alignment = Alignment(horizontal="right", vertical="center")

    # Col I: Like Rate = likes / alcance * 100
    c = ws3.cell(row=R, column=9)
    c.value = f"=IF(C{R}=0,0,D{R}/C{R}*100)"
    c.font = normal_font; c.fill = fill; c.border = thin_border()
    c.number_format = "0.00"; c.alignment = Alignment(horizontal="right", vertical="center")

    # Col J: Save Rate = guardados / alcance * 100
    c = ws3.cell(row=R, column=10)
    c.value = f"=IF(C{R}=0,0,F{R}/C{R}*100)"
    c.font = bold_font; c.fill = LAVENDER_FILL; c.border = thin_border()
    c.number_format = "0.00"; c.alignment = Alignment(horizontal="right", vertical="center")

    # Col K: Share Rate = compartidos / visualizaciones * 100  (referenciando Sheet 2)
    c = ws3.cell(row=R, column=11)
    c.value = f"=IF(C{R}=0,0,G{R}/C{R}*100)"
    c.font = normal_font; c.fill = fill; c.border = thin_border()
    c.number_format = "0.00"; c.alignment = Alignment(horizontal="right", vertical="center")

    # Col L: Follower Conv Rate = seguidores_ganados / alcance * 100
    seg_val = int(row["seguidores_ganados"])
    c = ws3.cell(row=R, column=12)
    c.value = f"=IF(C{R}=0,0,{seg_val}/C{R}*100)"
    c.font = normal_font; c.fill = fill; c.border = thin_border()
    c.number_format = "0.00"; c.alignment = Alignment(horizontal="right", vertical="center")

    ws3.row_dimensions[R].height = 18

# Summary row
SUM_ROW = len(reels) + 3
ws3.merge_cells(f"A{SUM_ROW}:B{SUM_ROW}")
ws3.cell(row=SUM_ROW, column=1, value="TOTAL / PROMEDIO").font = Font(bold=True, color=GOLD_C, size=11)
ws3.cell(row=SUM_ROW, column=1).fill = ACCENT_FILL
ws3.cell(row=SUM_ROW, column=1).border = thin_border()
for col in range(3, 13):
    ltr  = get_column_letter(col)
    cell = ws3.cell(row=SUM_ROW, column=col)
    first, last = 3, len(reels) + 2
    cell.value = f"=SUM({ltr}{first}:{ltr}{last})" if col <= 7 else f"=AVERAGE({ltr}{first}:{ltr}{last})"
    cell.number_format = "#,##0" if col <= 7 else "0.00"
    cell.font = Font(bold=True, color=GOLD_C, size=10)
    cell.fill = ACCENT_FILL; cell.border = thin_border()
    cell.alignment = Alignment(horizontal="right", vertical="center")
ws3.row_dimensions[SUM_ROW].height = 25

# ═════════════════════════════════════════════════════════════════════════════
# SHEET 4: MÉTRICAS AVANZADAS
# ═════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("📈 Métricas Avanzadas")
ws4.sheet_view.showGridLines = False

ws4.merge_cells("A1:H1")
ws4["A1"] = "📈 Métricas Avanzadas de Social Media — Instagram @aroaxinping"
ws4["A1"].font = title_font; ws4["A1"].fill = ACCENT_FILL
ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 35

# Glosario
explanations = [
    ("Follower Rate/1K",   "Seguidores ganados por cada 1.000 vistas. Mide la eficiencia de conversión de cada Reel."),
    ("Reach Rate %",       f"Alcance / {TOTAL_FOLLOWERS:,} seguidores × 100. % de tu comunidad que realmente ve el Reel."),
    ("Amplification Rate", f"Compartidos / {TOTAL_FOLLOWERS:,} seguidores × 100. Mide si tu comunidad te amplifica activamente."),
    ("Quality Score",      "Save Rate × Engagement Rate / 100. En Instagram los guardados son la señal de calidad más potente."),
    ("Passive vs Active",  "(Likes + Guardados) / (Comentarios + Compartidos). >5 = audiencia pasiva; <2 = muy activa e involucrada."),
]

ws4.merge_cells("A2:B2")
ws4["A2"] = "📖 Glosario de métricas"
ws4["A2"].font = subhead_font; ws4["A2"].fill = ACCENT_FILL
ws4["A2"].alignment = Alignment(horizontal="left", vertical="center")
ws4.row_dimensions[2].height = 22

ws4.column_dimensions["A"].width = 22
ws4.column_dimensions["B"].width = 72

for i, (name, desc) in enumerate(explanations, 3):
    fill = ROW_FILL_1 if i % 2 == 0 else ROW_FILL_2
    c1 = ws4.cell(row=i, column=1, value=name)
    c2 = ws4.cell(row=i, column=2, value=desc)
    c1.font = bold_font;   c1.fill = fill; c1.border = thin_border()
    c2.font = normal_font; c2.fill = fill; c2.border = thin_border()
    c1.alignment = Alignment(horizontal="left", vertical="center")
    c2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws4.row_dimensions[i].height = 20

# Tabla
DATA_ROW_START = len(explanations) + 4
ws4.merge_cells(f"A{DATA_ROW_START-1}:J{DATA_ROW_START-1}")
lbl = ws4.cell(row=DATA_ROW_START-1, column=1, value="📊 Datos por Reel")
lbl.font = subhead_font; lbl.fill = ACCENT_FILL
lbl.alignment = Alignment(horizontal="left", vertical="center")
ws4.row_dimensions[DATA_ROW_START-1].height = 22

col_names4  = ["#", "Descripción", "Vistas", "Alcance", "Seg.\nGanados",
               "Compartidos", "Likes", "Guardados",
               "Follower\nRate/1K", "Reach\nRate %",
               "Amplification\nRate %", "Quality\nScore", "Passive vs\nActive"]
col_widths4 = [4, 36, 12, 12, 9, 12, 10, 10, 14, 12, 16, 13, 14]

for i, (h, w) in enumerate(zip(col_names4, col_widths4), 1):
    ws4.column_dimensions[get_column_letter(i)].width = w
    c = ws4.cell(row=DATA_ROW_START, column=i, value=h); style_header(c)
ws4.row_dimensions[DATA_ROW_START].height = 32

for idx, (_, row) in enumerate(reels.iterrows(), DATA_ROW_START + 1):
    fill = ROW_FILL_1 if idx % 2 == 0 else ROW_FILL_2
    R = idx
    # Columns with raw data needed for formulas:
    # C=vistas, D=alcance, E=seg_ganados, F=compartidos, G=likes, H=guardados
    # We also need comentarios for Passive vs Active → store as a constant in formula
    comentarios_val = int(row["comentarios"])

    raw = [(1, idx - DATA_ROW_START),
           (2, row["descripcion_corta"][:40]),
           (3, int(row["visualizaciones"])),
           (4, int(row["alcance"])),
           (5, int(row["seguidores_ganados"])),
           (6, int(row["compartidos"])),
           (7, int(row["me_gustas"])),
           (8, int(row["guardados"]))]

    for col, val in raw:
        cell = ws4.cell(row=R, column=col, value=val)
        cell.font = normal_font; cell.fill = fill; cell.border = thin_border()
        cell.alignment = Alignment(horizontal="left" if col==2 else "right", vertical="center")
        if col in (3,4,5,6,7,8): cell.number_format = "#,##0"

    # Col I: Follower Rate/1K = seg_ganados / vistas * 1000
    c = ws4.cell(row=R, column=9)
    c.value = f"=IF(C{R}=0,0,E{R}/C{R}*1000)"
    c.font = bold_font; c.fill = LAVENDER_FILL; c.border = thin_border()
    c.number_format = "0.00"; c.alignment = Alignment(horizontal="right", vertical="center")

    # Col J: Reach Rate = alcance / total_followers * 100
    c = ws4.cell(row=R, column=10)
    c.value = f"=IF({TOTAL_FOLLOWERS}=0,0,D{R}/{TOTAL_FOLLOWERS}*100)"
    c.font = normal_font; c.fill = fill; c.border = thin_border()
    c.number_format = "0.00"; c.alignment = Alignment(horizontal="right", vertical="center")

    # Col K: Amplification Rate = compartidos / total_followers * 100
    c = ws4.cell(row=R, column=11)
    c.value = f"=IF({TOTAL_FOLLOWERS}=0,0,F{R}/{TOTAL_FOLLOWERS}*100)"
    c.font = normal_font; c.fill = fill; c.border = thin_border()
    c.number_format = "0.00"; c.alignment = Alignment(horizontal="right", vertical="center")

    # Col L: Quality Score = (guardados/alcance*100) * (likes+coments+guardados+compartidos)/alcance
    c = ws4.cell(row=R, column=12)
    c.value = f"=IF(D{R}=0,0,(H{R}/D{R}*100)*((G{R}+{comentarios_val}+H{R}+F{R})/D{R}*100)/100)"
    c.font = bold_font; c.fill = LAVENDER_FILL; c.border = thin_border()
    c.number_format = "0.00"; c.alignment = Alignment(horizontal="right", vertical="center")

    # Col M: Passive vs Active = (likes+guardados) / (comentarios+compartidos)
    c = ws4.cell(row=R, column=13)
    c.value = f'=IF(({comentarios_val}+F{R})=0,"N/A",(G{R}+H{R})/({comentarios_val}+F{R}))'
    c.font = normal_font; c.fill = fill; c.border = thin_border()
    c.number_format = "0.00"; c.alignment = Alignment(horizontal="right", vertical="center")

    ws4.row_dimensions[R].height = 18

# Summary
ADV_SUM = DATA_ROW_START + 1 + len(reels)
ws4.merge_cells(f"A{ADV_SUM}:B{ADV_SUM}")
ws4.cell(row=ADV_SUM, column=1, value="PROMEDIO").font = Font(bold=True, color=GOLD_C, size=11)
ws4.cell(row=ADV_SUM, column=1).fill = ACCENT_FILL
ws4.cell(row=ADV_SUM, column=1).border = thin_border()
first_d, last_d = DATA_ROW_START + 1, DATA_ROW_START + len(reels)
for col in range(3, 14):
    ltr  = get_column_letter(col)
    cell = ws4.cell(row=ADV_SUM, column=col)
    cell.value = f"=AVERAGE({ltr}{first_d}:{ltr}{last_d})"
    cell.number_format = "#,##0" if col <= 8 else "0.00"
    cell.font = Font(bold=True, color=GOLD_C, size=10)
    cell.fill = ACCENT_FILL; cell.border = thin_border()
    cell.alignment = Alignment(horizontal="right", vertical="center")
ws4.row_dimensions[ADV_SUM].height = 25

# ── Guardar ───────────────────────────────────────────────────────────────────
out_path = BASE / "instagram_analytics_aroaxinping.xlsx"
wb.save(out_path)
print(f"\n✅ Excel guardado en: {out_path}")
print("   Hojas: Overview · Reels Raw · Engagement Calc · Métricas Avanzadas")
